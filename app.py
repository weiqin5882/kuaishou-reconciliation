#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
快手订单对账与利润统计系统
基于 Flask 的网页版订单比对工具
"""

import os
import re
import uuid
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename

app = Flask(__name__)

# 配置
UPLOAD_FOLDER = 'uploads'
EXPORT_FOLDER = 'exports'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'et'}
MAX_CONTENT_LENGTH = 100 * 1024 * 1024  # 100MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXPORT_FOLDER'] = EXPORT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# 确保目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def clean_order_id(order_id):
    """清洗订单号：去除空格、特殊字符，只保留数字和字母"""
    if pd.isna(order_id):
        return ''
    order_id = str(order_id).strip()
    # 去除所有空格和特殊字符，只保留数字和字母
    order_id = re.sub(r'[^\w]', '', order_id)
    return order_id.upper()


def detect_columns(df, file_type='flow'):
    """自动检测列名"""
    columns = df.columns.tolist()
    result = {}
    
    # 订单号可能的列名
    order_patterns = ['订单号', '订单编号', '订单ID', 'order', 'order_id', '订单']
    # 商品名可能的列名
    product_patterns = ['商品', '商品名称', '商品名', 'product', '商品标题', '标题']
    # 实际结算金额可能的列名（订单流水表用）
    settlement_patterns = ['实际结算金额', '结算金额', '实付金额', '实际金额', '结算']
    # 销售额可能的列名（客服表用）
    amount_patterns = ['金额', '销售额', '实付金额', '支付金额', '总价', 'amount', 'price', '价格']
    # 成本可能的列名
    cost_patterns = ['成本', '成本价', '进货价', 'cost', '进价', '采购价']
    
    def find_column(patterns, columns):
        """查找匹配的列名"""
        for pattern in patterns:
            for col in columns:
                if pattern.lower() in col.lower():
                    return col
        return columns[0] if columns else None
    
    result['order_id'] = find_column(order_patterns, columns)
    result['product'] = find_column(product_patterns, columns)
    
    if file_type == 'flow':
        # 订单流水表：检测实际结算金额列
        result['settlement'] = find_column(settlement_patterns, columns)
        result['amount'] = result['settlement']  # 兼容
    else:
        # 客服表：检测销售额和成本列
        result['amount'] = find_column(amount_patterns, columns)
        result['cost'] = find_column(cost_patterns, columns)
    
    return result, columns


def process_flow_data(df, column_mapping):
    """处理订单流水数据（不需要状态过滤，都是有效订单）"""
    try:
        # 提取需要的列
        order_col = column_mapping.get('order_id')
        product_col = column_mapping.get('product')
        settlement_col = column_mapping.get('settlement') or column_mapping.get('amount')
        
        if not all([order_col, product_col, settlement_col]):
            raise ValueError("订单流水表缺少必要的列映射")
        
        # 创建新 DataFrame
        processed = pd.DataFrame()
        processed['订单号'] = df[order_col].apply(clean_order_id)
        processed['商品名称'] = df[product_col].astype(str)
        processed['实际结算金额'] = pd.to_numeric(df[settlement_col], errors='coerce').fillna(0)
        
        # 过滤空订单号
        processed = processed[processed['订单号'] != '']
        
        # 去重（保留第一条）
        processed = processed.drop_duplicates(subset=['订单号'], keep='first')
        
        return processed
    except Exception as e:
        raise Exception(f"处理订单流水数据时出错: {str(e)}")


def process_customer_data(df, column_mapping):
    """处理客服统计数据"""
    try:
        # 提取需要的列
        order_col = column_mapping.get('order_id')
        product_col = column_mapping.get('product')
        amount_col = column_mapping.get('amount')
        cost_col = column_mapping.get('cost')
        
        if not all([order_col, product_col, amount_col]):
            raise ValueError("客服表缺少必要的列映射")
        
        # 创建新 DataFrame
        processed = pd.DataFrame()
        processed['订单号'] = df[order_col].apply(clean_order_id)
        processed['商品名称'] = df[product_col].astype(str)
        processed['销售额'] = pd.to_numeric(df[amount_col], errors='coerce').fillna(0)
        
        # 成本列必须存在
        if cost_col and cost_col in df.columns:
            processed['成本'] = pd.to_numeric(df[cost_col], errors='coerce').fillna(0)
        else:
            processed['成本'] = 0
        
        # 过滤空订单号
        processed = processed[processed['订单号'] != '']
        
        # 检测重复的订单（保留所有重复记录用于展示）
        duplicates_mask = processed.duplicated(subset=['订单号'], keep=False)
        duplicates_df = processed[duplicates_mask].copy()
        duplicate_order_ids = duplicates_df['订单号'].unique().tolist()
        
        # 去重（保留第一条用于对比）
        processed = processed.drop_duplicates(subset=['订单号'], keep='first')
        
        return processed, duplicate_order_ids, duplicates_df
    except Exception as e:
        raise Exception(f"处理客服数据时出错: {str(e)}")


def compare_orders(flow_df, customer_df):
    """对比订单数据"""
    # 获取订单号集合
    flow_orders = set(flow_df['订单号'].tolist())
    customer_orders = set(customer_df['订单号'].tolist())
    
    # 匹配成功的订单
    matched_orders = flow_orders & customer_orders
    
    # 流水有但客服没有（漏单）
    missing_in_customer = flow_orders - customer_orders
    
    # 客服有但流水没有
    extra_in_customer = customer_orders - flow_orders
    
    return {
        'matched': matched_orders,
        'missing': missing_in_customer,
        'extra': extra_in_customer
    }


def generate_result_data(flow_df, customer_df, comparison_result, duplicate_orders, duplicates_df):
    """生成结果数据"""
    matched_orders = comparison_result['matched']
    missing_orders = comparison_result['missing']
    extra_orders = comparison_result['extra']
    
    result_rows = []
    
    # 1. 匹配成功的订单（利润 = 流水实际结算金额 - 客服成本）
    for order_id in matched_orders:
        flow_row = flow_df[flow_df['订单号'] == order_id].iloc[0]
        customer_row = customer_df[customer_df['订单号'] == order_id].iloc[0]
        
        settlement_amount = flow_row['实际结算金额']
        cost_amount = customer_row['成本']
        profit = settlement_amount - cost_amount
        
        # 检查是否在重复列表中
        is_duplicate = order_id in duplicate_orders
        remark = '【客服重复】' if is_duplicate else ''
        
        result_rows.append({
            '订单号': order_id,
            '商品名称': customer_row['商品名称'],
            '实际结算金额': settlement_amount,
            '成本': cost_amount,
            '利润': profit,
            '状态': '正常',
            '备注': remark
        })
    
    # 2. 流水有但客服没有的订单（漏单）
    for order_id in missing_orders:
        flow_row = flow_df[flow_df['订单号'] == order_id].iloc[0]
        result_rows.append({
            '订单号': order_id,
            '商品名称': flow_row['商品名称'],
            '实际结算金额': flow_row['实际结算金额'],
            '成本': 0,
            '利润': flow_row['实际结算金额'],
            '状态': '异常',
            '备注': '【客服漏记】'
        })
    
    # 3. 客服有但流水没有的订单
    for order_id in extra_orders:
        customer_row = customer_df[customer_df['订单号'] == order_id].iloc[0]
        result_rows.append({
            '订单号': order_id,
            '商品名称': customer_row['商品名称'],
            '实际结算金额': 0,
            '成本': customer_row['成本'],
            '利润': -customer_row['成本'],
            '状态': '异常',
            '备注': '【客服有流水没有】'
        })
    
    return pd.DataFrame(result_rows)


def generate_summary(result_df, duplicate_orders):
    """生成汇总统计"""
    normal_df = result_df[result_df['状态'] == '正常']
    abnormal_df = result_df[result_df['状态'] == '异常']
    duplicate_df = result_df[result_df['备注'].str.contains('重复', na=False)]
    
    summary = {
        'total_settlement': result_df['实际结算金额'].sum(),
        'total_cost': result_df['成本'].sum(),
        'total_profit': result_df['利润'].sum(),
        'total_orders': len(result_df),
        'normal_orders': len(normal_df),
        'abnormal_orders': len(abnormal_df),
        'missing_orders': len(result_df[result_df['备注'] == '【客服漏记】']),
        'extra_orders': len(result_df[result_df['备注'] == '【客服有流水没有】']),
        'duplicate_orders': len(duplicate_orders)
    }
    
    return summary


def export_to_excel(result_df, duplicate_df, summary, filename):
    """导出结果到 Excel（包含重复订单明细sheet）"""
    try:
        # 确保数值列是数字类型
        for col in ['实际结算金额', '成本', '利润']:
            if col in result_df.columns:
                result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0)
        
        # 重置索引并添加序号列
        result_df = result_df.reset_index(drop=True)
        result_df.insert(0, '序号', range(1, len(result_df) + 1))
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # ===== Sheet 1: 对账结果 =====
        ws = wb.active
        ws.title = "对账结果"
        
        # 设置列宽
        column_widths = {
            'A': 8,   # 序号
            'B': 25,  # 订单号
            'C': 40,  # 商品名称
            'D': 14,  # 实际结算金额
            'E': 12,  # 成本
            'F': 12,  # 利润
            'G': 12,  # 状态
            'H': 20   # 备注
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 定义样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        normal_font = Font(name='微软雅黑', size=10)
        red_font = Font(name='微软雅黑', size=10, color='FF0000')
        abnormal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        duplicate_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = ['序号', '订单号', '商品名称', '实际结算金额', '成本', '利润', '状态', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 写入数据
        col_order = ['序号', '订单号', '商品名称', '实际结算金额', '成本', '利润', '状态', '备注']
        
        for row_idx in range(len(result_df)):
            row_dict = result_df.iloc[row_idx].to_dict()
            
            is_abnormal = str(row_dict.get('状态', '')) == '异常'
            is_loss = float(row_dict.get('利润', 0)) < 0
            is_duplicate = '重复' in str(row_dict.get('备注', ''))
            
            for col_idx, col_name in enumerate(col_order, 1):
                value = row_dict.get(col_name, '')
                cell = ws.cell(row=row_idx + 2, column=col_idx)
                cell.border = thin_border
                
                if col_name == '序号':
                    cell.value = int(value) if value else row_idx + 1
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = normal_font
                elif col_name == '订单号':
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = normal_font
                elif col_name == '商品名称':
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = normal_font
                elif col_name in ['实际结算金额', '成本', '利润']:
                    cell.value = float(value) if value else 0
                    cell.number_format = '¥#,##0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if col_name == '利润' and is_loss:
                        cell.font = red_font
                    else:
                        cell.font = normal_font
                elif col_name == '状态':
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = normal_font
                elif col_name == '备注':
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = normal_font
                
                # 重复订单用浅红色高亮，异常用浅黄色
                if is_duplicate:
                    cell.fill = duplicate_fill
                elif is_abnormal and col_name not in ['实际结算金额', '成本', '利润']:
                    cell.fill = abnormal_fill
                elif is_abnormal and col_name == '利润' and not is_loss:
                    cell.fill = abnormal_fill
        
        # 添加汇总统计
        current_row = len(result_df) + 3
        
        summary_data = [
            ['汇总统计', '', '', '', '', '', '', ''],
            ['项目', '数值', '', '', '', '', '', ''],
            ['总结算金额', summary['total_settlement'], '', '', '', '', '', ''],
            ['总成本', summary['total_cost'], '', '', '', '', '', ''],
            ['总利润', summary['total_profit'], '', '', '', '', '', ''],
            ['订单总数', summary['total_orders'], '', '', '', '', '', ''],
            ['正常订单', summary['normal_orders'], '', '', '', '', '', ''],
            ['异常订单', summary['abnormal_orders'], '', '', '', '', '', ''],
            ['客服漏记', summary['missing_orders'], '', '', '', '', '', ''],
            ['客服有流水没有', summary['extra_orders'], '', '', '', '', '', ''],
            ['客服重复订单', summary['duplicate_orders'], '', '', '', '', '', '']
        ]
        
        summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        summary_font = Font(name='微软雅黑', size=10, bold=True)
        
        for row_data in summary_data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
                
                if current_row == len(result_df) + 3:
                    cell.fill = header_fill
                    cell.font = header_font
                elif current_row == len(result_df) + 4:
                    cell.fill = summary_fill
                    cell.font = summary_font
                else:
                    cell.font = normal_font
                    if col_idx == 2 and isinstance(value, (int, float)):
                        cell.number_format = '¥#,##0.00' if current_row <= len(result_df) + 7 else '#,##0'
            
            current_row += 1
        
        # 设置行高
        ws.row_dimensions[1].height = 25
        for row in range(2, current_row):
            ws.row_dimensions[row].height = 20
        
        # ===== Sheet 2: 客服重复订单明细 =====
        if not duplicate_df.empty:
            ws2 = wb.create_sheet("客服重复订单")
            dup_headers = ['订单号', '商品名称', '销售额', '成本']
            
            for col_idx, header in enumerate(dup_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            for row_idx, (_, row) in enumerate(duplicate_df.iterrows(), start=2):
                ws2.cell(row=row_idx, column=1, value=str(row.get('订单号', ''))).border = thin_border
                ws2.cell(row=row_idx, column=2, value=str(row.get('商品名称', ''))).border = thin_border
                ws2.cell(row=row_idx, column=3, value=float(row.get('销售额', 0))).border = thin_border
                ws2.cell(row=row_idx, column=3).number_format = '¥#,##0.00'
                ws2.cell(row=row_idx, column=4, value=float(row.get('成本', 0))).border = thin_border
                ws2.cell(row=row_idx, column=4).number_format = '¥#,##0.00'
            
            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 40
            ws2.column_dimensions['C'].width = 12
            ws2.column_dimensions['D'].width = 12
        
        # 保存文件
        filepath = os.path.join(EXPORT_FOLDER, filename)
        wb.save(filepath)
        return filepath
    except Exception as e:
        raise Exception(f"导出 Excel 时出错: {str(e)}")


# ============== API 路由 ==============

@app.route('/')
def index():
    """主页"""
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传文件并检测列"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'})
        
        file = request.files['file']
        file_type = request.form.get('type', 'flow')  # flow 或 customer
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': '不支持的文件格式，请上传 .xlsx 或 .xls 文件'})
        
        # 保存文件
        unique_id = str(uuid.uuid4())[:8]
        filename = f"{file_type}_{unique_id}_{secure_filename(file.filename)}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # 读取文件检测列
        try:
            if filename.endswith('.et'):
                return jsonify({
                    'success': False, 
                    'error': '检测到 WPS .et 格式，请另存为 .xlsx 格式后重新上传'
                })
            
            df = pd.read_excel(filepath, engine='openpyxl')
            
            if df.empty:
                return jsonify({'success': False, 'error': '文件为空'})
            
            # 自动检测列
            detected_columns, all_columns = detect_columns(df, file_type=file_type)
            
            return jsonify({
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'columns': all_columns,
                'detected': detected_columns,
                'row_count': len(df)
            })
            
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'success': False, 'error': f'读取文件失败: {str(e)}'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'上传失败: {str(e)}'})


@app.route('/api/compare', methods=['POST'])
def compare():
    """执行对比"""
    try:
        data = request.json
        
        flow_file = data.get('flow_file')
        customer_file = data.get('customer_file')
        flow_mapping = data.get('flow_mapping')
        customer_mapping = data.get('customer_mapping')
        
        if not all([flow_file, customer_file, flow_mapping, customer_mapping]):
            return jsonify({'success': False, 'error': '缺少必要的参数'})
        
        # 读取文件
        flow_df = pd.read_excel(flow_file, engine='openpyxl')
        customer_df = pd.read_excel(customer_file, engine='openpyxl')
        
        # 处理数据
        flow_processed = process_flow_data(flow_df, flow_mapping)
        customer_processed, duplicate_orders, duplicates_df = process_customer_data(customer_df, customer_mapping)
        
        # 对比订单
        comparison_result = compare_orders(flow_processed, customer_processed)
        
        # 生成结果
        result_df = generate_result_data(flow_processed, customer_processed, comparison_result, duplicate_orders, duplicates_df)
        
        # 生成汇总
        summary = generate_summary(result_df, duplicate_orders)
        
        # 转换为前端格式
        result_data = result_df.to_dict('records')
        duplicates_data = duplicates_df.to_dict('records') if not duplicates_df.empty else []
        
        return jsonify({
            'success': True,
            'data': result_data,
            'summary': summary,
            'duplicates': duplicate_orders,
            'duplicates_detail': duplicates_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/export', methods=['POST'])
def export():
    """导出 Excel"""
    try:
        data = request.json
        result_data = data.get('data', [])
        duplicates_data = data.get('duplicates_detail', [])
        summary = data.get('summary', {})
        
        if not result_data:
            return jsonify({'success': False, 'error': '没有数据可导出'})
        
        result_df = pd.DataFrame(result_data)
        duplicates_df = pd.DataFrame(duplicates_data) if duplicates_data else pd.DataFrame()
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"对账结果_{timestamp}.xlsx"
        
        # 导出
        filepath = export_to_excel(result_df, duplicates_df, summary, filename)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/api/download/{filename}'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/download/<filename>')
def download(filename):
    """下载文件"""
    try:
        filepath = os.path.join(EXPORT_FOLDER, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': '文件不存在'})
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/cleanup', methods=['POST'])
def cleanup():
    """清理临时文件"""
    try:
        data = request.json
        files = data.get('files', [])
        
        for filepath in files:
            if os.path.exists(filepath):
                os.remove(filepath)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    print("=" * 50)
    print("快手订单对账与利润统计系统")
    print("=" * 50)
    print("访问地址: http://localhost:5000")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=True)
